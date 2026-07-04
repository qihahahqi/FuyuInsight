#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
导入导出服务
"""

from typing import List, Dict, Optional
from io import BytesIO
from datetime import datetime, date
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExportService:
    """导入导出服务"""

    # 表头样式
    HEADER_FILL = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
    SUMMARY_FILL = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")

    # 资产类型配置
    ASSET_TYPES = {
        'market': {
            'name': '市价型产品',
            'types': ['stock', 'etf_index', 'etf_sector', 'fund', 'gold', 'silver'],
            'type_names': {'stock': '股票', 'etf_index': '宽基ETF', 'etf_sector': '行业ETF', 'fund': '基金', 'gold': '黄金', 'silver': '白银'}
        },
        'fixed_income': {
            'name': '固定收益类',
            'types': ['bank_deposit', 'bank_current', 'bank_wealth', 'treasury_bond', 'corporate_bond', 'money_fund'],
            'type_names': {'bank_deposit': '银行定期存款', 'bank_current': '银行活期存款', 'bank_wealth': '银行理财产品', 'treasury_bond': '国债', 'corporate_bond': '企业债', 'money_fund': '货币基金'}
        },
        'manual': {
            'name': '其他产品',
            'types': ['insurance', 'trust', 'other'],
            'type_names': {'insurance': '保险理财', 'trust': '信托产品', 'other': '其他'}
        }
    }

    # 产品大类中文名
    CATEGORY_NAMES = {'market': '市价型', 'fixed_income': '固定收益', 'manual': '其他'}

    # ============================================================
    #  持仓导出（增强版：增加业务字段 + 汇总统计 Sheet）
    # ============================================================
    def export_positions_to_excel(self, positions: List[Dict]) -> BytesIO:
        """导出持仓到 Excel，包含明细和汇总两个 Sheet"""
        wb = Workbook()
        # ---------- Sheet 1: 持仓明细 ----------
        ws = wb.active
        ws.title = "持仓明细"

        headers = [
            "账户", "代码", "名称", "产品大类", "类型",
            "数量", "成本价", "现价", "总成本", "市值",
            "盈亏金额", "收益率", "到期日", "风险等级", "分类", "备注"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.HEADER_ALIGNMENT

        for row, p in enumerate(positions, 2):
            cost = float(p.get('total_cost', 0) or 0)
            mv = float(p.get('market_value', 0) or 0)
            profit_amt = mv - cost if mv and cost else ''
            profit_rate = p.get('profit_rate')

            ws.cell(row=row, column=1, value=p.get('account_name') or '')
            ws.cell(row=row, column=2, value=p.get('symbol', ''))
            ws.cell(row=row, column=3, value=p.get('name', ''))
            ws.cell(row=row, column=4, value=self.CATEGORY_NAMES.get(p.get('product_category', ''), ''))
            ws.cell(row=row, column=5, value=self._get_asset_type_label(p.get('asset_type', '')))
            ws.cell(row=row, column=6, value=p.get('quantity', 0))
            ws.cell(row=row, column=7, value=p.get('cost_price', 0))
            ws.cell(row=row, column=8, value=p.get('current_price') or '')
            ws.cell(row=row, column=9, value=cost or '')
            ws.cell(row=row, column=10, value=mv or '')
            ws.cell(row=row, column=11, value=round(profit_amt, 2) if profit_amt != '' else '')
            ws.cell(row=row, column=12, value=f"{profit_rate*100:.2f}%" if profit_rate is not None else '')
            ws.cell(row=row, column=13, value=p.get('mature_date') or '')
            ws.cell(row=row, column=14, value=p.get('risk_level') or '')
            ws.cell(row=row, column=15, value=p.get('category') or '')
            ws.cell(row=row, column=16, value=p.get('notes') or '')

        # 列宽
        widths = [10, 12, 16, 10, 12, 10, 10, 10, 12, 12, 12, 10, 12, 10, 10, 20]
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = w

        # ---------- Sheet 2: 汇总统计 ----------
        ws2 = wb.create_sheet("汇总统计")

        total_cost = sum(float(p.get('total_cost', 0) or 0) for p in positions)
        total_mv = sum(float(p.get('market_value', 0) or 0) for p in positions)
        total_profit = total_mv - total_cost
        profit_pct = (total_profit / total_cost * 100) if total_cost > 0 else 0

        # 按产品大类统计
        by_category = {}
        for p in positions:
            cat = p.get('product_category', 'market')
            if cat not in by_category:
                by_category[cat] = {'cost': 0, 'mv': 0, 'count': 0}
            by_category[cat]['cost'] += float(p.get('total_cost', 0) or 0)
            by_category[cat]['mv'] += float(p.get('market_value', 0) or 0)
            by_category[cat]['count'] += 1

        # 按账户统计
        by_account = {}
        for p in positions:
            acct = p.get('account_name') or '默认账户'
            if acct not in by_account:
                by_account[acct] = {'cost': 0, 'mv': 0, 'count': 0}
            by_account[acct]['cost'] += float(p.get('total_cost', 0) or 0)
            by_account[acct]['mv'] += float(p.get('market_value', 0) or 0)
            by_account[acct]['count'] += 1

        row_idx = 1
        # 总览
        ws2.cell(row=row_idx, column=1, value="📊 持仓汇总统计").font = Font(bold=True, size=14)
        row_idx += 2

        overview = [
            ["持仓总数", f"{len(positions)} 条"],
            ["总成本", f"{total_cost:,.2f} 元"],
            ["总市值", f"{total_mv:,.2f} 元"],
            ["总盈亏", f"{total_profit:+,.2f} 元"],
            ["总收益率", f"{profit_pct:+.2f}%"],
            ["导出时间", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ]
        for item in overview:
            ws2.cell(row=row_idx, column=1, value=item[0]).font = Font(bold=True)
            ws2.cell(row=row_idx, column=2, value=item[1])
            if item[0] == "总盈亏" and total_profit >= 0:
                ws2.cell(row=row_idx, column=2).font = Font(color="22C55E")
            elif item[0] == "总盈亏" and total_profit < 0:
                ws2.cell(row=row_idx, column=2).font = Font(color="EF4444")
            row_idx += 1

        # 按产品大类
        row_idx += 1
        ws2.cell(row=row_idx, column=1, value="按产品大类").font = Font(bold=True, size=12)
        row_idx += 1
        for label in ["大类", "持仓数", "成本", "市值", "占比"]:
            c = ws2.cell(row=row_idx, column=["大类", "持仓数", "成本", "市值", "占比"].index(label) + 1, value=label)
            c.fill = self.HEADER_FILL
            c.font = self.HEADER_FONT
        row_idx += 1
        for cat, data in by_category.items():
            ws2.cell(row=row_idx, column=1, value=self.CATEGORY_NAMES.get(cat, cat))
            ws2.cell(row=row_idx, column=2, value=data['count'])
            ws2.cell(row=row_idx, column=3, value=f"{data['cost']:,.2f}")
            ws2.cell(row=row_idx, column=4, value=f"{data['mv']:,.2f}")
            ws2.cell(row=row_idx, column=5, value=f"{data['mv']/total_mv*100:.1f}%" if total_mv else '0%')
            row_idx += 1

        # 按账户
        if len(by_account) > 1:
            row_idx += 1
            ws2.cell(row=row_idx, column=1, value="按账户").font = Font(bold=True, size=12)
            row_idx += 1
            for label in ["账户", "持仓数", "成本", "市值", "占比"]:
                c = ws2.cell(row=row_idx, column=["账户", "持仓数", "成本", "市值", "占比"].index(label) + 1, value=label)
                c.fill = self.HEADER_FILL
                c.font = self.HEADER_FONT
            row_idx += 1
            for acct, data in by_account.items():
                ws2.cell(row=row_idx, column=1, value=acct)
                ws2.cell(row=row_idx, column=2, value=data['count'])
                ws2.cell(row=row_idx, column=3, value=f"{data['cost']:,.2f}")
                ws2.cell(row=row_idx, column=4, value=f"{data['mv']:,.2f}")
                ws2.cell(row=row_idx, column=5, value=f"{data['mv']/total_mv*100:.1f}%" if total_mv else '0%')
                row_idx += 1

        ws2.column_dimensions['A'].width = 16
        ws2.column_dimensions['B'].width = 18
        ws2.column_dimensions['C'].width = 16
        ws2.column_dimensions['D'].width = 16
        ws2.column_dimensions['E'].width = 10

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # ============================================================
    #  交易导出（增强版：增加标的名称、账户、手续费）
    # ============================================================
    def export_trades_to_excel(self, trades: List[Dict]) -> BytesIO:
        """导出交易记录到 Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "交易记录"

        headers = ["日期", "账户", "代码", "名称", "类型", "数量", "价格", "金额", "手续费", "交易理由", "备注"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.HEADER_ALIGNMENT

        for row, t in enumerate(trades, 2):
            ws.cell(row=row, column=1, value=t.get('trade_date', ''))
            ws.cell(row=row, column=2, value=t.get('account_name') or '')
            ws.cell(row=row, column=3, value=t.get('symbol', ''))
            ws.cell(row=row, column=4, value=t.get('position_name') or '')
            ws.cell(row=row, column=5, value='买入' if t.get('trade_type') == 'buy' else '卖出')
            ws.cell(row=row, column=6, value=t.get('quantity', 0))
            ws.cell(row=row, column=7, value=t.get('price', 0))
            ws.cell(row=row, column=8, value=t.get('amount', 0))
            ws.cell(row=row, column=9, value=t.get('commission') or '')
            ws.cell(row=row, column=10, value=t.get('reason') or '')
            ws.cell(row=row, column=11, value=t.get('notes') or '')

        widths = [12, 10, 12, 16, 8, 10, 10, 12, 10, 16, 20]
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = w

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # ============================================================
    #  持仓导入（增强列映射，修复起息日/到期日 Bug）
    # ============================================================
    def import_positions_from_excel(self, file_stream) -> List[Dict]:
        """从 Excel 导入持仓"""
        all_sheets = pd.read_excel(file_stream, sheet_name=None)

        # 列名映射（支持带*的必填标记）
        column_map = {
            '代码': 'symbol', '代码*': 'symbol',
            '名称': 'name', '名称*': 'name',
            '类型': 'asset_type', '类型*': 'asset_type',
            '数量': 'quantity', '数量*': 'quantity',
            '成本价': 'cost_price', '成本价*': 'cost_price',
            '成本': 'cost_price', '成本*': 'cost_price',
            '本金金额': 'cost_price', '本金金额*': 'cost_price',
            '现价': 'current_price',
            '总成本': 'total_cost',
            '市值': 'market_value',
            '收益率': 'profit_rate',
            '分类': 'category',
            '备注': 'notes',
            '产品大类': 'product_category',
            '到期日': 'mature_date',
            '起息日': 'start_date',
            '风险等级': 'risk_level',
            '预期收益率': 'expected_return',
            '预期收益率%': 'expected_return',
            '年化利率%': 'expected_return',
            '账户': 'account_name',
        }

        type_map = {
            '宽基ETF': 'etf_index', '行业ETF': 'etf_sector', '基金': 'fund', '股票': 'stock',
            '黄金': 'gold', '白银': 'silver',
            '银行定期存款': 'bank_deposit', '银行活期存款': 'bank_current', '银行理财产品': 'bank_wealth',
            '国债': 'treasury_bond', '企业债': 'corporate_bond', '货币基金': 'money_fund',
            '保险理财': 'insurance', '信托产品': 'trust', '其他': 'other'
        }

        results = []
        for sheet_name, df in all_sheets.items():
            if '说明' in sheet_name or '使用' in sheet_name or '统计' in sheet_name or len(df) < 2:
                continue
            df = df.rename(columns=column_map)
            if 'symbol' not in df.columns or 'name' not in df.columns:
                continue

            for _, row in df.iterrows():
                if pd.isna(row.get('symbol')) or pd.isna(row.get('name')):
                    continue

                asset_type = row.get('asset_type', '')
                if asset_type in type_map:
                    asset_type = type_map[asset_type]
                elif not asset_type:
                    asset_type = 'etf_index'

                # 处理收益率字符串
                profit_rate = row.get('profit_rate')
                if isinstance(profit_rate, str):
                    try:
                        profit_rate = float(profit_rate.replace('%', '')) / 100
                    except:
                        profit_rate = None

                # 处理预期收益率（支持百分号格式）
                expected_return = row.get('expected_return')
                if isinstance(expected_return, str):
                    try:
                        expected_return = float(expected_return.replace('%', '')) / 100
                    except:
                        expected_return = None
                elif isinstance(expected_return, (int, float)) and not pd.isna(expected_return):
                    # 如果数值 >1 视为百分比值（如 3.5 = 3.5%）
                    if expected_return > 1:
                        expected_return = expected_return / 100

                # 处理起息日/到期日
                start_date = row.get('start_date')
                if isinstance(start_date, str):
                    start_date = start_date.strip()
                elif hasattr(start_date, 'strftime'):
                    start_date = start_date.strftime('%Y-%m-%d')
                mature_date = row.get('mature_date')
                if isinstance(mature_date, str):
                    mature_date = mature_date.strip()
                elif hasattr(mature_date, 'strftime'):
                    mature_date = mature_date.strftime('%Y-%m-%d')

                try:
                    quantity = float(row.get('quantity', 0)) if not pd.isna(row.get('quantity')) else 0
                    cost_price = float(row.get('cost_price', 0)) if not pd.isna(row.get('cost_price')) else 0
                except:
                    continue

                if quantity <= 0 or cost_price <= 0:
                    continue

                symbol_raw = row.get('symbol', '')
                symbol = self._format_symbol(symbol_raw, asset_type)

                results.append({
                    'symbol': symbol,
                    'name': str(row.get('name', '')).strip(),
                    'asset_type': asset_type,
                    'quantity': quantity,
                    'cost_price': cost_price,
                    'current_price': float(row.get('current_price')) if not pd.isna(row.get('current_price')) else None,
                    'total_cost': float(row.get('total_cost', 0)) if not pd.isna(row.get('total_cost')) else 0,
                    'market_value': float(row.get('market_value')) if not pd.isna(row.get('market_value')) else None,
                    'profit_rate': profit_rate,
                    'category': str(row.get('category', '')).strip() if not pd.isna(row.get('category')) else None,
                    'notes': str(row.get('notes', '')).strip() if not pd.isna(row.get('notes')) else None,
                    'product_category': str(row.get('product_category', '')).strip() if not pd.isna(row.get('product_category')) else None,
                    'start_date': start_date if not pd.isna(row.get('start_date')) else None,
                    'mature_date': mature_date if not pd.isna(row.get('mature_date')) else None,
                    'risk_level': str(row.get('risk_level', '')).strip() if not pd.isna(row.get('risk_level')) else None,
                    'expected_return': expected_return,
                    'account_name': str(row.get('account_name', '')).strip() if not pd.isna(row.get('account_name')) else None,
                })

        return results

    def import_positions_from_csv(self, file_stream) -> List[Dict]:
        """从 CSV 导入持仓"""
        df = pd.read_csv(file_stream)
        return self._parse_dataframe(df)

    def _parse_dataframe(self, df: pd.DataFrame) -> List[Dict]:
        """解析 DataFrame"""
        results = []
        for _, row in df.iterrows():
            results.append({
                'symbol': str(row.get('symbol', '')).strip(),
                'name': str(row.get('name', '')).strip(),
                'asset_type': row.get('asset_type', 'etf_index'),
                'quantity': float(row.get('quantity', 0)),
                'cost_price': float(row.get('cost_price', 0)),
                'current_price': float(row.get('current_price')) if pd.notna(row.get('current_price')) else None,
                'category': row.get('category'),
                'notes': row.get('notes')
            })
        return results

    def _get_asset_type_label(self, asset_type: str) -> str:
        """获取资产类型标签"""
        labels = {
            'etf_index': '宽基ETF', 'etf_sector': '行业ETF', 'fund': '基金', 'stock': '股票',
            'gold': '黄金', 'silver': '白银',
            'bank_deposit': '银行定期存款', 'bank_current': '银行活期存款', 'bank_wealth': '银行理财产品',
            'treasury_bond': '国债', 'corporate_bond': '企业债', 'money_fund': '货币基金',
            'insurance': '保险理财', 'trust': '信托产品', 'other': '其他'
        }
        return labels.get(asset_type, asset_type)

    def _format_symbol(self, symbol_raw, asset_type: str) -> str:
        """格式化代码，基金类型补齐6位"""
        if pd.isna(symbol_raw):
            return ''
        symbol = str(symbol_raw).strip()
        fund_types = ['fund', 'etf_index', 'etf_sector']
        if asset_type in fund_types:
            if symbol.isdigit() and len(symbol) < 6:
                symbol = symbol.zfill(6)
        return symbol

    # ============================================================
    #  持仓导入模板（修复：起息日/到期日 分开）
    # ============================================================
    def export_positions_template(self, category: str = None) -> BytesIO:
        """生成持仓导入模板"""
        wb = Workbook()

        # ---------- 导入说明 ----------
        ws_overview = wb.active
        ws_overview.title = "导入说明"

        overview_content = [
            ["📋 持仓导入模板使用说明"],
            [],
            ["一、必填字段（标记 * 的列不能为空）"],
            ["  代码：市价型产品必填（股票/ETF/基金），固定收益类可留空"],
            ["  名称：产品名称"],
            ["  数量：持有数量或份额"],
            ["  成本价/本金金额：买入价格或本金金额"],
            [],
            ["二、资产类型对照表"],
            ["  市价型：股票、宽基ETF、行业ETF、基金、黄金、白银"],
            ["  固定收益：银行定期存款、银行活期存款、银行理财产品、国债、企业债、货币基金"],
            ["  其他：保险理财、信托产品、其他"],
            [],
            ["三、导入注意事项"],
            ["  1. 请勿修改表头名称，表头带 * 号的为必填列"],
            ["  2. 必填字段不能为空"],
            ["  3. 日期格式：YYYY-MM-DD（如 2024-01-15）"],
            ["  4. 收益率填写百分比数值，如 3.5 表示 3.5%"],
            ["  5. 已存在的代码+账户组合会自动跳过，不会重复导入"],
            ["  6. 起息日和到期日分别填写，请勿合并"],
            [],
            ["四、各类型模板工作表"],
            ["  请切换到对应的工作表填写数据："]
        ]

        for row_idx, row_data in enumerate(overview_content, 1):
            if row_data:
                ws_overview.cell(row=row_idx, column=1, value=row_data[0])
                if row_idx == 1:
                    ws_overview.cell(row=row_idx, column=1).font = Font(bold=True, size=14)
                elif any(row_data[0].startswith(p) for p in ['一、', '二、', '三、', '四、']):
                    ws_overview.cell(row=row_idx, column=1).font = Font(bold=True, size=12)

        ws_overview.column_dimensions['A'].width = 65

        # ---------- 按分类创建工作表 ----------
        categories_to_create = [category] if category else ['market', 'fixed_income', 'manual']

        for cat in categories_to_create:
            if cat not in self.ASSET_TYPES:
                continue

            cat_config = self.ASSET_TYPES[cat]
            ws = wb.create_sheet(title=cat_config['name'])

            if cat == 'market':
                headers = ["代码*", "名称*", "类型*", "数量*", "成本价*", "现价", "分类", "账户", "备注"]
                example_data = [
                    ["000001", "平安银行", "股票", 1000, 10.5, 11.2, "core", "主账户", "观察仓"],
                    ["510300", "沪深300ETF", "宽基ETF", 5000, 4.0, 4.2, "core", "主账户", "定投标的"],
                    ["159915", "创业板ETF", "宽基ETF", 3000, 2.5, "", "satellite", "主账户", ""],
                    ["515030", "新能源ETF", "行业ETF", 2000, 3.0, 2.8, "aggressive", "主账户", "风险仓"],
                ]
            elif cat == 'fixed_income':
                # 🔧 修复：起息日和到期日分成两列
                headers = ["代码", "名称*", "类型*", "本金金额*", "年化利率%", "起息日", "到期日", "风险等级", "账户", "备注"]
                example_data = [
                    ["", "工商银行一年定存", "银行定期存款", 100000, 2.5, "2024-01-15", "2025-01-15", "R1", "主账户", "一年期"],
                    ["", "招商银行理财", "银行理财产品", 50000, 3.8, "2024-02-01", "2024-08-01", "R2", "主账户", "半年期"],
                    ["", "2024年国债", "国债", 20000, 2.8, "2024-01-01", "2027-01-01", "R1", "主账户", "三年期"],
                    ["", "余额宝", "货币基金", 30000, 1.8, "", "", "R1", "主账户", "活期理财"],
                ]
            else:  # manual
                headers = ["代码", "名称*", "类型*", "本金金额*", "预期收益率%", "起息日", "到期日", "风险等级", "账户", "备注"]
                example_data = [
                    ["", "平安保险理财", "保险理财", 30000, 3.5, "2024-03-01", "2025-03-01", "R2", "主账户", ""],
                    ["", "信托产品A", "信托产品", 100000, 5.0, "2024-01-01", "2026-01-01", "R4", "主账户", "两年期"],
                    ["", "其他投资", "其他", 5000, "", "", "", "", "主账户", "自定义记录"],
                ]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT
                cell.alignment = self.HEADER_ALIGNMENT

            for row_idx, row_data in enumerate(example_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # 类型说明
            type_row = len(example_data) + 3
            ws.cell(row=type_row, column=1, value="可选类型：").font = Font(bold=True)
            ws.cell(row=type_row, column=2, value=", ".join(cat_config['type_names'].values()))

            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 16

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # ============================================================
    #  交易导入模板（增强：增加标的名称、手续费）
    # ============================================================
    def export_trades_template(self) -> BytesIO:
        """生成交易记录导入模板"""
        wb = Workbook()
        ws = wb.active
        ws.title = "交易记录导入模板"

        headers = ["日期*", "代码*", "名称", "类型*", "数量*", "价格*", "手续费", "账户", "交易理由", "备注"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.HEADER_ALIGNMENT

        example_data = [
            ["2024-01-15", "510300", "沪深300ETF", "买入", 1000, 4.000, 5.00, "主账户", "低估建仓", "定投"],
            ["2024-02-20", "510300", "沪深300ETF", "买入", 500, 3.900, 5.00, "主账户", "加仓", ""],
            ["2024-03-10", "159915", "创业板ETF", "卖出", 1000, 2.800, 5.00, "主账户", "止盈", "获利了结"],
        ]

        for row_idx, row_data in enumerate(example_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # ---------- 字段说明 ----------
        ws_desc = wb.create_sheet("字段说明")
        descriptions = [
            ["字段名", "是否必填", "说明", "示例值"],
            ["日期", "是", "交易日期，格式：YYYY-MM-DD", "2024-01-15"],
            ["代码", "是", "标的代码，如 510300；基金代码不足6位自动补齐", "510300"],
            ["名称", "否", "标的名称（选填，方便辨识）", "沪深300ETF"],
            ["类型", "是", "买入 或 卖出", "买入"],
            ["数量", "是", "交易数量（股/份）", "1000"],
            ["价格", "是", "交易单价", "4.000"],
            ["手续费", "否", "交易佣金/手续费（元）", "5.00"],
            ["账户", "否", "交易账户名称", "主账户"],
            ["交易理由", "否", "如：低估建仓/止盈/止损/定投", "低估建仓"],
            ["备注", "否", "其他备注信息", "定投"],
        ]

        for row_idx, row_data in enumerate(descriptions, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_desc.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.fill = self.HEADER_FILL
                    cell.font = self.HEADER_FONT

        widths = [12, 10, 35, 16]
        for i, w in enumerate(widths):
            ws_desc.column_dimensions[get_column_letter(i + 1)].width = w

        main_widths = [12, 12, 16, 8, 10, 10, 10, 10, 16, 20]
        for i, w in enumerate(main_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = w

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # ============================================================
    #  交易导入（增强：支持手续费、账户、名称）
    # ============================================================
    def import_trades_from_excel(self, file_stream) -> List[Dict]:
        """从 Excel 导入交易记录"""
        df = pd.read_excel(file_stream, sheet_name=0)

        column_map = {
            '日期': 'trade_date', '日期*': 'trade_date',
            '代码': 'symbol', '代码*': 'symbol',
            '名称': 'position_name',
            '类型': 'trade_type', '类型*': 'trade_type',
            '数量': 'quantity', '数量*': 'quantity',
            '价格': 'price', '价格*': 'price',
            '金额': 'amount',
            '手续费': 'commission',
            '佣金': 'commission',
            '账户': 'account_name',
            '理由': 'reason', '交易理由': 'reason',
            '备注': 'notes'
        }

        df = df.rename(columns=column_map)

        results = []
        for _, row in df.iterrows():
            if pd.isna(row.get('trade_date')) or pd.isna(row.get('symbol')):
                continue

            trade_type = str(row.get('trade_type', '')).strip()
            if trade_type == '买入':
                trade_type = 'buy'
            elif trade_type == '卖出':
                trade_type = 'sell'
            elif trade_type not in ['buy', 'sell']:
                continue

            trade_date = row.get('trade_date')
            if isinstance(trade_date, str):
                trade_date = trade_date.strip()
            elif hasattr(trade_date, 'strftime'):
                trade_date = trade_date.strftime('%Y-%m-%d')

            symbol_raw = row.get('symbol', '')
            symbol = str(symbol_raw).strip()
            if symbol.isdigit() and len(symbol) < 6:
                symbol = symbol.zfill(6)

            commission = row.get('commission')
            if not pd.isna(commission):
                try:
                    commission = float(commission)
                except:
                    commission = None
            else:
                commission = None

            results.append({
                'trade_date': trade_date,
                'symbol': symbol,
                'trade_type': trade_type,
                'quantity': int(row.get('quantity', 0)) if not pd.isna(row.get('quantity')) else 0,
                'price': float(row.get('price', 0)) if not pd.isna(row.get('price')) else 0,
                'amount': float(row.get('amount', 0)) if not pd.isna(row.get('amount')) else None,
                'commission': commission,
                'position_name': str(row.get('position_name', '')).strip() if not pd.isna(row.get('position_name')) else None,
                'account_name': str(row.get('account_name', '')).strip() if not pd.isna(row.get('account_name')) else None,
                'reason': str(row.get('reason', '')).strip() if not pd.isna(row.get('reason')) else None,
                'notes': str(row.get('notes', '')).strip() if not pd.isna(row.get('notes')) else None
            })

        return results
